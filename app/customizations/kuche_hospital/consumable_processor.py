from __future__ import annotations
import time
import threading
import re
import os
from datetime import datetime
from typing import Callable, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

from .element_loader import ElementLoader


class ConsumableProcessor:
    """
    耗材采购处理器 - v6.0 (智能等待版)
    
    核心改进：
    1. 自动模式/手动模式切换（手动模式用于调试）
    2. 智能等待：基于页面状态验证，非固定时间等待
    3. 优化速度：去除不必要的延时
    """
    
    # 等待超时配置
    TIMEOUT_DIALOG = 10.0       # 弹窗出现超时
    TIMEOUT_ELEMENT = 5.0       # 元素就绪超时
    TIMEOUT_QUERY = 15.0        # 查询结果超时
    
    # 弹窗根路径
    DIALOG_XPATH = '//div[contains(@class,"el-dialog__wrapper") and not(contains(@style,"display: none"))]'
    
    def __init__(
        self, 
        browser_tab: Any,
        progress_callback: Callable[[str], None] | None = None,
        confirm_callback: Callable[[str], bool] | None = None,
        debug_mode: bool = False,  # 手动模式（带断点确认）
        auto_mode: bool = True     # 自动模式（无断点，全自动）
    ):
        self.tab = browser_tab
        self.progress_callback = progress_callback or print
        self.confirm_callback = confirm_callback
        self.debug_mode = debug_mode and not auto_mode  # 只有非自动模式才启用调试
        self.auto_mode = auto_mode
        self._processed_codes: set = set()
        self._stop_requested = False
        self._pause_requested = False
        self._pause_event = threading.Event()
        self._pause_event.set()
        
        self.loader = ElementLoader()
        self._current_frame = None  
    
    def _log(self, msg: str):
        if self.progress_callback:
            self.progress_callback(msg)

    def stop(self):
        self._stop_requested = True
        self._pause_event.set()
    
    def pause(self):
        self._pause_requested = True
        self._pause_event.clear()
    
    def resume(self):
        self._pause_requested = False
        self._pause_event.set()
    
    def _check_stop(self) -> bool:
        """检查是否需要停止（优雅停止：完成当前行后再停止）"""
        return self._stop_requested
    
    def _check_pause(self):
        self._pause_event.wait()
    
    def _wait_confirm(self, step_name: str) -> bool:
        if not self.debug_mode or not self.confirm_callback:
            return True
        self._log(f"🔵 [断点] {step_name}")
        result = self.confirm_callback(step_name)
        if not result:
            self._log(f"❌ 用户在 [{step_name}] 选择终止")
            self._stop_requested = True
            return False
        self._log(f"✅ 用户确认 [{step_name}] 正确")
        return True

    def _get_target(self):
        return self._current_frame or self.tab

    def _find_element_recursive(self, root: Any, selectors: list[str], timeout: float = 1.0, depth: int = 0) -> tuple[Any | None, Any]:
        """递归在所有 frame 中查找元素 (深度优先)"""
        for selector in selectors:
            try:
                elem = root.ele(selector, timeout=0.3) 
                if elem:
                    return elem, root
            except:
                continue
        if depth > 5: return None, root
        try:
            iframes = root.eles('tag:iframe')
            for iframe in iframes:
                res, owner = self._find_element_recursive(iframe, selectors, timeout, depth + 1)
                if res: return res, owner
        except: pass
        return None, root

    def process(self, excel_data: pd.DataFrame, code_column: str = '医保码') -> dict[str, Any]:
        """执行流程 (v5.0-价格审核版)"""
        self._stop_requested = False
        self._log("🚀 [v5.0-价格审核版] 开始耗材采购自动化处理...")
        
        if code_column not in excel_data.columns:
            possible = ['医保码', 'C码', '耗材代码', '产品代码', '医用耗材代码']
            for name in possible:
                if name in excel_data.columns:
                    code_column = name
                    break
            else:
                self._log(f"❌ 找不到代码列")
                return {'success': False, 'msg': '找不到代码列'}
        
        # 查找价格列
        price_column = None
        possible_price_cols = ['医院采购价', '单价(元)', '单价（元）', '单价', '价格']
        for col in possible_price_cols:
            if col in excel_data.columns:
                price_column = col
                self._log(f"💰 找到价格列: {col}")
                break
        if not price_column:
            self._log("⚠️ 未找到价格列，将跳过价格审核")
        
        # 保存 Excel 数据供价格查询（按医保码分组，取第一条的价格）
        self._excel_prices = {}
        if price_column:
            for code_val in excel_data[code_column].unique():
                rows = excel_data[excel_data[code_column] == code_val]
                if not rows.empty:
                    price = rows.iloc[0][price_column]
                    try:
                        self._excel_prices[str(code_val).strip()] = float(price)
                    except:
                        pass
            self._log(f"📊 已加载 {len(self._excel_prices)} 个产品的价格信息")

        counts = excel_data[code_column].value_counts().to_dict()
        total_unique = len(counts)
        self._log(f"📊 统计完成：共 {total_unique} 个唯一耗材代码")
        
        if not self._wait_confirm(f"数据统计完成，共 {total_unique} 个 C 码，准备开始自动录入"):
            return {'success': False, 'msg': '用户中止'}

        success_count = 0
        skipped_count = 0
        price_mismatch_count = 0  # 价格不一致跳过计数
        multi_result_count = 0    # 多条结果跳过计数
        need_open_dialog = True
        
        # 报告数据收集
        self._report_rows = []  # 每条记录: {状态, 医保码, 产品名称, 表格价格, 网页价格, 生产厂家, 数量}
        
        # 【新增】保存原始Excel数据和代码列名，用于生成基于原表的报告
        self._original_excel_data = excel_data.copy()
        self._code_column = code_column
        self._code_status = {}  # {医保码: 状态} 用于标记每个代码的处理结果
        
        # 保存产品名称和生产厂家信息（从Excel读取）
        self._excel_product_info = {}
        name_col = None
        manufacturer_col = None
        for col in excel_data.columns:
            if '产品名称' in col or '物资名称' in col or '耗材名称' in col:
                name_col = col
            if '生产厂家' in col or '厂家' in col or '生产企业' in col:
                manufacturer_col = col
        
        if name_col or manufacturer_col:
            for code_val in excel_data[code_column].unique():
                rows = excel_data[excel_data[code_column] == code_val]
                if not rows.empty:
                    self._excel_product_info[str(code_val).strip()] = {
                        '产品名称': rows.iloc[0][name_col] if name_col else '',
                        '生产厂家': rows.iloc[0][manufacturer_col] if manufacturer_col else ''
                    }
        
        mode_label = "自动模式" if self.auto_mode else "手动调试模式"
        self._log(f"🔧 运行模式：{mode_label}")

        for i, (code_val, count) in enumerate(counts.items(), 1):
            # 优雅停止：在开始新一条前检查（不是处理中间）
            if self._check_stop():
                self._log("⏹️ 用户终止，已完成当前行后停止")
                break
            self._check_pause()
            code = str(code_val).strip()
            self._log(f"📝 [{i}/{total_unique}] 处理 {code}（数量：{count}）...")
            
            # 1. 只有需要时才点击添加产品
            if need_open_dialog:
                if not self._click_add_product():
                    self._log(f"   ⚠️ 无法打开弹窗，跳过此条")
                    continue
                # 智能等待：验证弹窗已打开
                if not self._wait_dialog_open():
                    self._log(f"   ⚠️ 弹窗未正确打开")
                    continue
                need_open_dialog = False
            if not self._wait_confirm(f"已打开弹窗，是否正确？"): break
            
            # 2. 输入代码（智能等待输入框就绪）
            if not self._fill_code_input(code): continue
            if not self._wait_confirm(f"已输入代码 [{code}]"): break
            
            # 3. 点击查询（智能等待查询结果）
            if not self._click_query(): continue
            result_status, row_count = self._wait_for_results()
            
            # 获取产品信息用于报告
            product_info = self._excel_product_info.get(code, {})
            excel_price = self._excel_prices.get(code)
            
            if result_status == 'empty':
                self._log(f"   ➡️ 查询无结果，继续下一个")
                skipped_count += 1
                self._code_status[code] = '查无信息'
                continue
            elif result_status != 'data':
                continue
            
            # 3.3 多行结果检测：超过1行时跳过并记录
            if row_count > 1:
                self._log(f"   ⚠️ 查询返回 {row_count} 条结果，需要手动处理，跳过")
                multi_result_count += 1
                # 记录到报告
                self._report_rows.append({
                    '状态': f'多条结果({row_count}条)',
                    '医保码': code,
                    '产品名称': product_info.get('产品名称', ''),
                    '表格价格': excel_price or '',
                    '网页价格': '',
                    '生产厂家': product_info.get('生产厂家', ''),
                    '数量': count,
                    '备注': f'查询返回{row_count}条结果，可能价格/厂家/型号不同，需手动处理'
                })
                self._code_status[code] = f'多条结果({row_count}条)'
                continue
            
            # 3.5 价格审核：对比网页挂网价和Excel单价
            web_price = None
            
            if excel_price is not None:
                web_price = self._get_web_price()
                if web_price is not None:
                    if excel_price != web_price:  # 完全相等才录入
                        self._log(f"   ⚠️ 价格不一致！Excel: {excel_price} vs 网页: {web_price}，跳过")
                        price_mismatch_count += 1
                        # 记录到报告
                        self._report_rows.append({
                            '状态': '价格不一致',
                            '医保码': code,
                            '产品名称': product_info.get('产品名称', ''),
                            '表格价格': excel_price,
                            '网页价格': web_price,
                            '生产厂家': product_info.get('生产厂家', ''),
                            '数量': count
                        })
                        self._code_status[code] = '价格不一致'
                        continue
                    else:
                        self._log(f"   ✅ 价格一致: {excel_price}")
            
            if not self._wait_confirm(f"已查询 [{code}]"): break
            
            # 4. 填入数量（多行时填第一行）
            if not self._fill_quantity(int(count)): continue
            if not self._wait_confirm(f"已填入采购数量 [{count}]"): break
            
            # 5. 点击保存（智能等待弹窗关闭）
            if not self._click_save():
                self._log(f"   ⚠️ 保存失败，跳过此条")
                continue
            
            # 智能等待：验证弹窗已关闭
            if not self._wait_dialog_closed():
                self._log(f"   ⚠️ 弹窗未正确关闭")
            
            success_count += 1
            self._log(f"   ✅ [{i}/{total_unique}] 完成")
            need_open_dialog = True
            
            # 记录成功到报告
            self._report_rows.append({
                '状态': '已完成',
                '医保码': code,
                '产品名称': product_info.get('产品名称', ''),
                '表格价格': excel_price or '',
                '网页价格': web_price or '',
                '生产厂家': product_info.get('生产厂家', ''),
                '数量': count
            })
            self._code_status[code] = '已完成'
        
        # 报告数据已收集，可通过 export_report() 导出
        report_count = len(self._report_rows)
        
        self._log(f"🏁 处理完成：成功 {success_count}，跳过 {skipped_count}，价格不一致 {price_mismatch_count}，共 {total_unique}")
        if report_count > 0:
            self._log(f"📊 已记录 {report_count} 条数据，请点击[导出报告]保存")
        
        self._log(f"🏁 处理完成：成功 {success_count}，跳过 {skipped_count}，其中：价格不一致 {price_mismatch_count}，多条结果 {multi_result_count}")
        return {
            'success': success_count > 0, 
            'count': success_count, 
            'skipped': skipped_count, 
            'price_mismatch': price_mismatch_count, 
            'multi_result': multi_result_count,
            'report_count': report_count
        }

    def has_report_data(self) -> bool:
        """检查是否有报告数据可导出"""
        return hasattr(self, '_report_rows') and len(self._report_rows) > 0
    
    def get_report_count(self) -> int:
        """获取报告数据条数"""
        return len(self._report_rows) if hasattr(self, '_report_rows') else 0
    
    def export_report(self, filepath: str) -> bool:
        """
        导出Excel报告到指定路径（基于原表，带颜色标记）
        
        Args:
            filepath: 保存文件路径
            
        Returns:
            bool: 是否成功
        """
        if not hasattr(self, '_original_excel_data') or self._original_excel_data is None:
            self._log("⚠️ 无原始数据可导出")
            return False
        
        try:
            from copy import copy
            
            wb = Workbook()
            ws = wb.active
            ws.title = "采购处理报告"
            
            # 定义样式
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            success_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # 淡绿色（已完成）
            mismatch_fill = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")  # 橘黄色（价格不一致）
            multi_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")  # 淡黄色（多条结果）
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            df = self._original_excel_data
            code_col = self._code_column
            
            # 获取原始列名
            original_columns = list(df.columns)
            # 在第一列插入"录入状态"列
            all_columns = ['录入状态'] + original_columns
            
            # 写表头
            for col_idx, col_name in enumerate(all_columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=col_name)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
                cell.border = thin_border
            
            # 写数据行
            for row_idx, (_, row) in enumerate(df.iterrows(), 2):
                # 获取该行的医保码
                code_value = str(row.get(code_col, '')).strip()
                # 查找该代码的状态
                status = self._code_status.get(code_value, '')
                
                # 第一列：录入状态
                cell = ws.cell(row=row_idx, column=1, value=status)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center')
                
                # 根据状态设置整行背景色
                row_fill = None
                if status == '已完成':
                    row_fill = success_fill
                elif status == '价格不一致':
                    row_fill = mismatch_fill
                elif '多条结果' in status:
                    row_fill = multi_fill
                elif status == '查无信息':
                    row_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # 灰色
                
                if row_fill:
                    cell.fill = row_fill
                
                # 原始数据列
                for col_idx, col_name in enumerate(original_columns, 2):
                    value = row.get(col_name, '')
                    # 处理 NaN 值
                    if pd.isna(value):
                        value = ''
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = thin_border
                    if row_fill:
                        cell.fill = row_fill
            
            # 自动调整列宽（粗略估算）
            ws.column_dimensions['A'].width = 12  # 录入状态列
            for i, col_name in enumerate(original_columns, 2):
                col_letter = chr(64 + i) if i <= 26 else f"{chr(64 + i // 26)}{chr(64 + i % 26)}"
                try:
                    from openpyxl.utils import get_column_letter
                    col_letter = get_column_letter(i)
                    ws.column_dimensions[col_letter].width = min(max(len(str(col_name)) + 2, 10), 30)
                except:
                    pass
            
            # 保存文件
            wb.save(filepath)
            
            # 统计
            completed = sum(1 for s in self._code_status.values() if s == '已完成')
            total = len(df)
            self._log(f"📊 报告已保存: {filepath}")
            self._log(f"📊 统计: 总行数 {total}，已录入 {completed}，其他状态 {len(self._code_status) - completed}")
            return True
            
        except Exception as e:
            self._log(f"⚠️ 导出报告失败: {e}")
            import traceback
            traceback.print_exc()
            return False

    def _wait_dialog_open(self, timeout: float = None) -> bool:
        """智能等待：验证弹窗已打开"""
        timeout = timeout or self.TIMEOUT_DIALOG
        target = self._get_target()
        start = time.time()
        while time.time() - start < timeout:
            dialog = target.ele(f'xpath:{self.DIALOG_XPATH}', timeout=0.3)
            if dialog:
                # 验证弹窗内有内容（如表格或表单）
                content = dialog.ele('xpath:.//div[contains(@class,"el-dialog__body")]', timeout=0.2)
                if content:
                    return True
            time.sleep(0.1)
        return False

    def _wait_dialog_closed(self, timeout: float = 5.0) -> bool:
        """智能等待：验证弹窗已关闭"""
        target = self._get_target()
        start = time.time()
        while time.time() - start < timeout:
            dialog = target.ele(f'xpath:{self.DIALOG_XPATH}', timeout=0.2)
            if not dialog:
                return True
            time.sleep(0.1)
        return False
    
    def _get_web_price(self) -> float | None:
        """
        从查询结果表格中提取医院采购价
        
        Returns:
            float: 医院采购价，获取失败返回 None
        """
        try:
            target = self._get_target()
            
            # 策略1：通过表头找到"医院采购价"列的索引，然后获取第一行的对应单元格
            # Element UI 表格结构：表头在 thead，数据在 tbody
            
            # 先找表头中包含"医院采购价"的列
            header_selectors = [
                f'xpath:{self.DIALOG_XPATH}//thead//th[contains(.,\"医院采购价\")]',
                f'xpath:{self.DIALOG_XPATH}//div[contains(@class,\"el-table__header\")]//th[contains(.,\"医院采购价\")]',
            ]
            
            header_cell = None
            for sel in header_selectors:
                header_cell = target.ele(sel, timeout=0.5)
                if header_cell:
                    break
            
            if header_cell:
                # 获取列索引（通过 JS 获取 cellIndex）
                col_index = header_cell.run_js('return this.cellIndex;')
                if col_index is not None:
                    # 获取第一行数据的对应单元格
                    # 注意：使用主表体区域，不使用固定列区域（避免重复）
                    cell_selector = f'xpath:{self.DIALOG_XPATH}//div[contains(@class,\"el-table__body-wrapper\") and not(ancestor::div[contains(@class,\"el-table__fixed\")])]//tr[contains(@class,\"el-table__row\")][1]//td[{col_index + 1}]'
                    cell = target.ele(cell_selector, timeout=0.5)
                    if cell:
                        price_text = cell.text.strip()
                        # 清理价格文本（去除货币符号等）
                        price_text = re.sub(r'[^\d.]', '', price_text)
                        if price_text:
                            price = float(price_text)
                            self._log(f"   💰 网页医院采购价: {price}")
                            return price
            
            # 策略2：直接搜索包含数字格式的单元格（兜底）
            self._log("   ⚠️ 未能定位医院采购价列")
            return None
            
        except Exception as e:
            self._log(f"   ⚠️ 获取网页价格失败: {e}")
            return None

    def _click_add_product(self) -> bool:
        """点击添加产品 - v6.1 增强重试版"""
        try:
            self._log("🔘 点击【添加产品】...")
            
            # 页面可能需要时间加载，先等待一下
            time.sleep(1.5)
            
            selectors = [
                'xpath://button[contains(@class,"el-button--success")][contains(.,"添加产品")]',
                'xpath://button[contains(.,"添加产品")]',
                'xpath://button[.//span[contains(text(),"添加产品")]]',
                'xpath://div[contains(@class,"elian-tool_right")]//button[contains(.,"添加产品")]',
                'text=添加产品',
                'text:添加产品'
            ]
            
            # 最多尝试5次，每次间隔1秒
            for attempt in range(5):
                self._log(f"   🔄 第 {attempt+1} 次尝试...")
                btn, owner = self._find_element_recursive(self.tab, selectors, timeout=3.0)
                if btn:
                    self._current_frame = owner
                    try:
                        btn.click()
                    except:
                        btn.run_js('this.click();')
                    self._log(f"   ✅ 已点击")
                    return True
                time.sleep(1.0)
            
            self._log("   ❌ 无法定位【添加产品】按钮，请确保页面已完全加载")
            return False
        except Exception as e:
            self._log(f"   ❌ 异常: {e}")
            return False

    def _fill_code_input(self, code: str) -> bool:
        """填写代码"""
        try:
            target = self._get_target()
            selectors = [
                f'xpath:{self.DIALOG_XPATH}//label[contains(.,"医用耗材代码")]/following-sibling::div//input',
                f'xpath:{self.DIALOG_XPATH}//input[@placeholder[contains(.,"代码")]]',
                f'xpath:{self.DIALOG_XPATH}//div[contains(@class,"el-form-item")]//input'
            ]
            for s in selectors:
                ele = target.ele(s, timeout=1.5)
                if ele:
                    ele.click()
                    ele.clear()
                    ele.input(code)
                    time.sleep(0.2)
                    ele.run_js('this.dispatchEvent(new KeyboardEvent("keydown", {key: "Enter", keyCode: 13, bubbles: true}));')
                    self._log(f"   📝 输入: {code} (并尝试回车)")
                    return True
            self._log(f"   ❌ 找不到代码输入框")
            return False
        except Exception as e:
            self._log(f"   ❌ 代码输入异常: {e}")
            return False

    def _click_query(self) -> bool:
        """点击查询"""
        try:
            target = self._get_target()
            selectors = [
                f'xpath:{self.DIALOG_XPATH}//button[contains(.,"查询")]',
                f'xpath:{self.DIALOG_XPATH}//button[.//span[contains(.,"查询")]]',
                f'xpath:{self.DIALOG_XPATH}//div[contains(@class,"el-form-item")]//button'
            ]
            for i, s in enumerate(selectors):
                btn = target.ele(s, timeout=1.5)
                if btn:
                    # 直接在按钮元素上调用 JS，避免跨 frame 问题
                    btn.run_js('this.scrollIntoView({block: "center"}); this.click();')
                    self._log(f"   🔍 已穿透触发查询 (v4.6)")
                    return True
            self._log("   ❌ 找不到【查询】按钮")
            return False
        except Exception as e:
            self._log(f"   ❌ 查询按钮异常: {e}")
            return False

    def _wait_for_results(self, timeout: float = 10.0) -> str:
        """等待结果，返回: 'data'=有数据, 'empty'=无数据, ''=超时/错误"""
        try:
            target = self._get_target()
            start = time.time()
            self._log(f"   ⏳ 等待数据加载...")
            
            # 先等一下让页面渲染
            time.sleep(1.0)
            
            while time.time() - start < timeout:
                # 只检查主表格区域（不含固定列区域）的数据行
                # Element UI 的固定列会复制 tr，导致重复计数
                main_body_rows = target.eles(
                    f'xpath:{self.DIALOG_XPATH}//div[contains(@class,"el-table__body-wrapper") and not(ancestor::div[contains(@class,"el-table__fixed")])]//tr[contains(@class,"el-table__row")]'
                )
                if main_body_rows and len(main_body_rows) > 0:
                    row_count = len(main_body_rows)
                    self._log(f"   ✨ 已检测到 {row_count} 条数据")
                    return ('data', row_count)
                
                # 检查 Element UI 的空状态元素
                empty_block = target.ele(f'xpath:{self.DIALOG_XPATH}//div[contains(@class,"el-table__empty-block")]', timeout=0.3)
                if empty_block:
                    self._log("   ⚠️ 检测到空数据状态，将跳过此条")
                    return ('empty', 0)
                
                # 检查常见的空数据提示文字
                page_text = target.html or ""
                empty_patterns = ["暂无数据", "无数据", "没有数据", "无匹配", "未找到", "No Data"]
                for pattern in empty_patterns:
                    if pattern in page_text:
                        self._log(f"   ⚠️ 检测到'{pattern}'，将跳过此条")
                        return ('empty', 0)
                
                time.sleep(0.5)
            
            self._log("   ❌ 等待查询结果超时")
            return ('', 0)
        except Exception as e:
            self._log(f"   ❌ 等待结果异常: {e}")
            return ('', 0)

    def _fill_quantity(self, quantity: int) -> bool:
        """填入数量 - 终极版 v5.4 (采购数量=第二个输入框)"""
        try:
            target = self._get_target()
            
            self._log("   📊 开始定位采购数量输入框（右侧固定列的第二个）...")
            
            # 策略1：直接找 maxlength=12 的输入框（采购数量独有特征）
            sel1 = f'xpath:{self.DIALOG_XPATH}//div[contains(@class,"el-table__fixed-right")]//input[@maxlength="12"]'
            inputs = target.eles(sel1)
            self._log(f"      maxlength=12 找到 {len(inputs)} 个")
            
            for inp in inputs:
                try:
                    checks = inp.run_js('return {inForm: this.closest(".el-form-item") !== null, visible: this.offsetWidth > 0};')
                    if checks.get('inForm') or not checks.get('visible'):
                        continue
                    
                    inp.run_js('this.scrollIntoView({block: "center"});')
                    time.sleep(0.2)
                    inp.click()
                    time.sleep(0.1)
                    inp.clear()
                    inp.input(str(quantity))
                    time.sleep(0.2)
                    self._log(f"   🎯 maxlength=12 定位成功，录入 {quantity}")
                    return True
                except:
                    continue
            
            # 策略2：找右侧固定列的所有可见输入框，跳过第一个（配送企业），选第二个
            sel2 = f'xpath:{self.DIALOG_XPATH}//div[contains(@class,"el-table__fixed-right")]//tr[contains(@class,"el-table__row")]//input'
            all_inputs = target.eles(sel2)
            self._log(f"      右侧固定列行内找到 {len(all_inputs)} 个输入框")
            
            visible_inputs = []
            for inp in all_inputs:
                try:
                    is_visible = inp.run_js('return this.offsetWidth > 0 && this.offsetHeight > 0;')
                    if is_visible:
                        visible_inputs.append(inp)
                except:
                    continue
            
            self._log(f"      其中可见的有 {len(visible_inputs)} 个")
            
            # 采购数量是第二个可见输入框（配送企业是第一个）
            if len(visible_inputs) >= 2:
                inp = visible_inputs[1]  # 取第二个
                inp.run_js('this.scrollIntoView({block: "center"});')
                time.sleep(0.2)
                inp.click()
                time.sleep(0.1)
                inp.clear()
                inp.input(str(quantity))
                time.sleep(0.2)
                self._log(f"   🎯 第二个输入框定位成功，录入 {quantity}")
                return True
            elif len(visible_inputs) == 1:
                # 只有一个的话就用那个
                inp = visible_inputs[0]
                inp.run_js('this.scrollIntoView({block: "center"});')
                time.sleep(0.2)
                inp.click()
                time.sleep(0.1)
                inp.clear()
                inp.input(str(quantity))
                time.sleep(0.2)
                self._log(f"   🎯 唯一输入框定位成功，录入 {quantity}")
                return True
            
            self._log("   ❌ 所有策略均失败")
            return False
        except Exception as e:
            self._log(f"   ❌ 填写数量异常: {e}")
            return False
            
            # 最后兜底：找所有弹窗内的 input，排除搜索区域
            self._log("   📊 执行兜底搜索...")
            all_inputs = target.eles(f'xpath:{self.DIALOG_XPATH}//input')
            for inp in all_inputs:
                try:
                    # 排除搜索区（在表头上方的）
                    rect = inp.run_js('return {y: this.getBoundingClientRect().top, inForm: this.closest(".el-form-item") !== null, visible: this.offsetWidth > 0};')
                    
                    if rect.get('inForm'):
                        continue
                    if not rect.get('visible'):
                        continue
                    if rect.get('y', 0) < 200:  # 假设搜索区在上方
                        continue
                    
                    placeholder = inp.attr('placeholder') or ''
                    if '请输入' in placeholder or '代码' in placeholder:
                        continue
                    
                    inp.click()
                    time.sleep(0.1)
                    inp.clear()
                    inp.input(str(quantity))
                    time.sleep(0.2)
                    self._log(f"   🎯 兜底策略成功，录入 {quantity}")
                    return True
                except:
                    continue
            
            self._log("   ❌ 所有策略均失败，无法定位采购数量输入框")
            return False
        except Exception as e:
            self._log(f"   ❌ 填写数量异常: {e}")
            return False

    def _click_save(self) -> bool:
        """点击保存"""
        try:
            target = self._get_target()
            self._log("   💾 点击保存...")
            selectors = [
                # 精确匹配用户提供的按钮特征
                f'xpath:{self.DIALOG_XPATH}//button[contains(@class,"el-button--primary")][contains(.,"保")]',
                f'xpath:{self.DIALOG_XPATH}//button[span[contains(text(),"保")]]',
                f'xpath:{self.DIALOG_XPATH}//button[contains(.,"保存")]',
                'xpath://button[contains(@class,"el-button--primary")][contains(.,"保")]',
                'text=保 存',
                'text=保存'
            ]
            btn, _ = self._find_element_recursive(self.tab, selectors, timeout=2.0)
            if btn:
                btn.run_js('this.click();')  # JS 点击更可靠
                time.sleep(1.5)  # 等待弹窗关闭
                self._log("   ✅ 已点击保存，弹窗已关闭")
                return True
            self._log("   ❌ 找不到保存按钮")
            return False
        except Exception as e:
            self._log(f"   ❌ 保存异常: {e}")
            return False

    def _close_dialog(self) -> bool:
        """关闭弹窗（用于查询无结果时）"""
        try:
            target = self._get_target()
            # 尝试点击关闭按钮 (X)
            close_selectors = [
                f'xpath:{self.DIALOG_XPATH}//button[contains(@class,"el-dialog__headerbtn")]',
                f'xpath:{self.DIALOG_XPATH}//i[contains(@class,"el-dialog__close")]',
                f'xpath:{self.DIALOG_XPATH}//button[contains(@class,"close")]',
            ]
            for sel in close_selectors:
                btn = target.ele(sel, timeout=0.5)
                if btn:
                    btn.run_js('this.click();')
                    time.sleep(0.5)
                    self._log("   ❎ 已关闭弹窗")
                    return True
            # 兜底：按 Escape 键
            target.run_js('document.dispatchEvent(new KeyboardEvent("keydown", {key: "Escape", keyCode: 27}));')
            time.sleep(0.5)
            self._log("   ❎ 已发送 Escape 关闭弹窗")
            return True
        except:
            return False
